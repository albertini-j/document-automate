"""
Document Control Automation Script (docctl.py)

Assumptions and safe defaults:
- Required fields in transmittal logs: TRANSMITTAL NUMBER, DATE, ITEM, DOCUMENT NUMBER 1, TITLE, VERSION.
- HEADER matching is case-insensitive; leading/trailing whitespace and trailing semicolons are ignored.
- DATE values are parsed with dateutil.parser if they are not already datetime/date objects; invalid dates cause rejection.
- VERSION values are treated as alphanumeric strings; whitespace is trimmed but formatting is preserved for storage.
- Document-to-file association uses case-insensitive substring matching on DOCUMENT NUMBER 1 within filenames (excluding the transmittal log itself).
- document_list.xlsx always reflects the latest accepted transmittal for each DOCUMENT NUMBER 1 (chronological override, no version sequencing logic).
- Versions already present in transmittal_database.xlsx for a DOCUMENT NUMBER 1 cannot be resubmitted; such transmittals are rejected.
- If destination folders already contain a transmittal with the same name, a numeric suffix (-1, -2, ...) is added to avoid overwriting.
- The Current Files directory is synced to the latest accepted transmittal by deleting older matching files and copying in the newly accepted files.
- All rows within a transmittal are validated to aggregate every error before deciding acceptance; remaining pending transmittals continue processing even after encountering errors.
"""
from __future__ import annotations

import argparse
import logging
import shutil
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Dict, List

from dateutil import parser as date_parser
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

EXPECTED_HEADERS = [
    "TRANSMITTAL NUMBER",
    "TRANSMITTAL NAME",
    "DATE",
    "ITEM",
    "DOCUMENT NUMBER 1",
    "DOCUMENT NUMBER 2",
    "TITLE",
    "VERSION",
    "DOCUMENT STATE",
    "ISSUE OBJECTIVE",
    "HOLD LIST",
    "ISSUED BY",
    "ISSUED TO",
]
DATABASE_HEADERS = EXPECTED_HEADERS + ["FILENAMES"]
REQUIRED_FIELDS = {
    "TRANSMITTAL NUMBER",
    "DATE",
    "ITEM",
    "DOCUMENT NUMBER 1",
    "TITLE",
    "VERSION",
}


@dataclass
class TransmittalRow:
    raw: Dict[str, object]
    normalized_doc: str
    normalized_version: str
    filenames: List[str]


def normalize_header(value: str) -> str:
    return value.strip().rstrip(";").upper()


def normalize_version(value: object) -> str:
    if value is None or str(value).strip() == "":
        raise ValueError("VERSION is empty")
    return str(value).strip()


def normalize_doc_number(value: object) -> str:
    if value is None or str(value).strip() == "":
        raise ValueError("DOCUMENT NUMBER 1 is empty")
    return str(value).strip().lower()


def coerce_date(value: object) -> object:
    if isinstance(value, (datetime, date)):
        return value
    if value is None:
        raise ValueError("DATE is empty")
    text = str(value).strip()
    if not text:
        raise ValueError("DATE is empty")
    try:
        return date_parser.parse(text)
    except (ValueError, TypeError) as exc:
        raise ValueError(f"Invalid DATE '{value}'") from exc


def read_sheet_rows(sheet: Worksheet) -> List[Dict[str, object]]:
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    if any(h is None for h in headers):
        raise ValueError("Header row contains empty cells")

    normalized_map = {normalize_header(h): h for h in headers}
    for expected in EXPECTED_HEADERS:
        if expected not in normalized_map:
            raise ValueError(f"Missing required column '{expected}' in header")

    rows: List[Dict[str, object]] = []
    for row in sheet.iter_rows(min_row=2):
        if all(cell.value in (None, "") for cell in row):
            continue
        record: Dict[str, object] = {}
        for cell, header in zip(row, headers):
            record[normalize_header(header)] = cell.value
        rows.append(record)
    return rows


def find_matching_files(transmittal_dir: Path, doc_number: str, log_file: Path) -> List[str]:
    matches: List[str] = []
    for entry in transmittal_dir.iterdir():
        if entry == log_file or entry.is_dir():
            continue
        if doc_number in entry.name.lower():
            matches.append(entry.name)
    return sorted(matches)


def ensure_workbook(path: Path, headers: List[str]) -> Worksheet:
    if path.exists():
        wb = load_workbook(path)
        sheet = wb.active
        existing_headers = [normalize_header(c.value or "") for c in next(sheet.iter_rows(min_row=1, max_row=1))]
        normalized_headers = [normalize_header(h) for h in headers]
        if existing_headers != normalized_headers:
            raise ValueError(f"Header mismatch in {path}")
        return sheet
    wb = Workbook()
    sheet = wb.active
    for col, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col, value=header)
    wb.save(path)
    return sheet


def load_existing_versions(db_path: Path) -> set[tuple[str, str]]:
    if not db_path.exists():
        return set()
    workbook = load_workbook(db_path)
    sheet = workbook.active
    headers = [normalize_header(c.value or "") for c in next(sheet.iter_rows(min_row=1, max_row=1))]
    header_map = {h: idx for idx, h in enumerate(headers)}
    required = {"DOCUMENT NUMBER 1", "VERSION"}
    if not required.issubset(header_map):
        raise ValueError(f"Header mismatch in {db_path}")
    seen: set[tuple[str, str]] = set()
    for row in sheet.iter_rows(min_row=2):
        doc_val = row[header_map["DOCUMENT NUMBER 1"]].value
        ver_val = row[header_map["VERSION"]].value
        if doc_val in (None, "") or ver_val in (None, ""):
            continue
        seen.add((normalize_doc_number(doc_val), normalize_version(ver_val).lower()))
    return seen


def append_to_database(db_path: Path, rows: List[TransmittalRow]) -> None:
    sheet = ensure_workbook(db_path, DATABASE_HEADERS)
    wb = sheet.parent
    for row in rows:
        values = [row.raw.get(h) for h in EXPECTED_HEADERS]
        values.append("; ".join(row.filenames))
        sheet.append(values)
    wb.save(db_path)


def update_document_list(doc_list_path: Path, rows: List[TransmittalRow]) -> None:
    sheet = ensure_workbook(doc_list_path, EXPECTED_HEADERS)
    wb = sheet.parent

    existing: Dict[str, Dict[str, object]] = {}
    for row in sheet.iter_rows(min_row=2):
        if all(cell.value in (None, "") for cell in row):
            continue
        record = {normalize_header(h.value): cell.value for h, cell in zip(sheet[1], row)}
        doc_key = normalize_doc_number(record["DOCUMENT NUMBER 1"])
        existing[doc_key] = record

    for row in rows:
        existing[row.normalized_doc] = row.raw

    sheet.delete_rows(2, sheet.max_row)
    for record in existing.values():
        sheet.append([record.get(h) for h in EXPECTED_HEADERS])
    wb.save(doc_list_path)


def move_transmittal(src: Path, dest_root: Path) -> Path:
    dest_root.mkdir(parents=True, exist_ok=True)
    candidate = dest_root / src.name
    counter = 1
    while candidate.exists():
        candidate = dest_root / f"{src.name}-{counter}"
        counter += 1
    shutil.move(str(src), candidate)
    return candidate


def sync_current_files(current_dir: Path, transmittal_dir: Path, rows: List[TransmittalRow], logger: logging.Logger) -> None:
    current_dir.mkdir(parents=True, exist_ok=True)

    processed_docs: set[str] = set()
    for row in rows:
        doc_key = row.normalized_doc
        if doc_key not in processed_docs:
            for existing in current_dir.iterdir():
                if existing.is_file() and doc_key in existing.name.lower():
                    try:
                        existing.unlink()
                        logger.info("Removed outdated current file %s for document %s", existing.name, row.raw["DOCUMENT NUMBER 1"])
                    except Exception as exc:  # pragma: no cover - defensive logging only
                        logger.warning("Failed to remove %s: %s", existing, exc)
            processed_docs.add(doc_key)

        for fname in row.filenames:
            src = transmittal_dir / fname
            if not src.exists():
                logger.warning("Listed file %s missing in transmittal %s", fname, transmittal_dir.name)
                continue
            dest = current_dir / fname
            shutil.copy2(src, dest)
            logger.info("Copied %s to Current Files for document %s", fname, row.raw["DOCUMENT NUMBER 1"])


def process_transmittal(transmittal_dir: Path, project_paths: Dict[str, Path], logger: logging.Logger) -> None:
    log_file = transmittal_dir / f"{transmittal_dir.name}.xlsx"
    if not log_file.exists():
        logger.error("Transmittal %s rejected: missing log %s", transmittal_dir.name, log_file.name)
        move_transmittal(transmittal_dir, project_paths["rejected"])
        return

    existing_versions = load_existing_versions(project_paths["database"])

    try:
        workbook = load_workbook(log_file)
        sheet = workbook.active
        raw_rows = read_sheet_rows(sheet)
        processed_rows: List[TransmittalRow] = []
        errors: List[str] = []
        for idx, raw in enumerate(raw_rows, start=2):
            try:
                for field in REQUIRED_FIELDS:
                    value = raw.get(field)
                    if value is None or str(value).strip() == "":
                        raise ValueError(f"Required field '{field}' is empty")
                normalized_doc = normalize_doc_number(raw["DOCUMENT NUMBER 1"])
                normalized_version = normalize_version(raw["VERSION"])
                version_key = (normalized_doc, normalized_version.lower())
                if version_key in existing_versions:
                    raise ValueError(
                        f"Duplicate version '{raw['VERSION']}' for document '{raw['DOCUMENT NUMBER 1']}' already submitted"
                    )
                raw["DATE"] = coerce_date(raw["DATE"])
                filenames = find_matching_files(transmittal_dir, normalized_doc, log_file)
                processed_rows.append(
                    TransmittalRow(
                        raw=raw,
                        normalized_doc=normalized_doc,
                        normalized_version=normalized_version,
                        filenames=filenames,
                    )
                )
            except Exception as exc:  # pragma: no cover - defensive aggregation for reporting
                errors.append(f"Row {idx}: {exc}")
                continue

        if errors:
            for msg in errors:
                logger.error("Transmittal %s validation error - %s", transmittal_dir.name, msg)
        if not processed_rows:
            raise ValueError("Transmittal contains no valid rows")
        if errors:
            raise ValueError("Transmittal rejected due to validation errors")
    except Exception as exc:
        logger.exception("Transmittal %s rejected: %s", transmittal_dir.name, exc)
        move_transmittal(transmittal_dir, project_paths["rejected"])
        return

    append_to_database(project_paths["database"], processed_rows)
    update_document_list(project_paths["doc_list"], processed_rows)
    sync_current_files(project_paths["current"], transmittal_dir, processed_rows, logger)
    moved_path = move_transmittal(transmittal_dir, project_paths["accepted"])
    logger.info("Transmittal %s accepted and moved to %s", transmittal_dir.name, moved_path)


def configure_logging(log_dir: Path) -> logging.Logger:
    log_dir.mkdir(parents=True, exist_ok=True)
    log_file = log_dir / "docctl.log"
    logger = logging.getLogger("docctl")
    logger.setLevel(logging.INFO)
    formatter = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")

    file_handler = logging.FileHandler(log_file)
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)

    console_handler = logging.StreamHandler()
    console_handler.setFormatter(formatter)
    logger.addHandler(console_handler)

    return logger


def ensure_project_paths(root: Path) -> Dict[str, Path]:
    paths = {
        "current": root / "Current Files",
        "pending": root / "Pending Transmittals",
        "accepted": root / "Accepted Transmittals",
        "rejected": root / "Rejected Transmittals",
        "reports": root / "Reports",
        "logs": root / "Logs",
    }
    for key, path in paths.items():
        path.mkdir(parents=True, exist_ok=True)
    paths["database"] = paths["reports"] / "transmittal_database.xlsx"
    paths["doc_list"] = paths["reports"] / "document_list.xlsx"
    return paths


def process_project(root: Path) -> None:
    project_paths = ensure_project_paths(root)
    logger = configure_logging(project_paths["logs"])

    pending_dir = project_paths["pending"]
    transmittals = [p for p in pending_dir.iterdir() if p.is_dir()]
    if not transmittals:
        logger.info("No pending transmittals found in %s", pending_dir)
        return

    for transmittal in sorted(transmittals):
        logger.info("Processing transmittal %s", transmittal.name)
        process_transmittal(transmittal, project_paths, logger)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Automate engineering document control workflows.")
    parser.add_argument(
        "--project-root",
        type=Path,
        default=Path.cwd(),
        help="Root directory of the project containing the required subfolders.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    process_project(args.project_root)


if __name__ == "__main__":
    main()
